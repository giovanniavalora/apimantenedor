##### Auth #####
from django.contrib.auth import authenticate
from django.contrib.auth.hashers import make_password
from django.contrib.auth.signals import user_logged_in
from django.core.exceptions import ObjectDoesNotExist

from rest_framework import status
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.decorators import api_view, permission_classes, authentication_classes

from django.conf import settings
################


from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import viewsets
from .serializers import *
from .models import *
from django.db.models import Count, Sum, IntegerField
from django.db.models.functions import Cast


from django.utils import timezone
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

@api_view(['GET'])
# @authentication_classes([])
# @permission_classes([])
def exportar_a_xlsx(request,start,end):

    if (not Voucher.objects.filter(fecha__range=(start,end)).exists()):
        print("\nno hay vouchers")
        res={}
        res = {'request': False, 'error': 'No existen registros para el rango especificado'}
        return Response(res,status=status.HTTP_204_NO_CONTENT)

    response = HttpResponse(
        # content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        content_type='application/vnd.ms-excel',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-reporte.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )

    voucher_queryset = Voucher.objects.filter(fecha__range=(start,end))
    print('voucher_queryset: ', voucher_queryset)

    
    # camiones_norm = Voucher.objects.annotate(as_integer=Cast('capacidad_camion',IntegerField()))

    # camion_queryset = Voucher.objects.filter(fecha__range=(start,end)).filter(available=True) \
    #     .values('available','nombre_subcontratista','patente_camion','marca_camion','modelo_camion',
    #     'color_camion','unidad_medida','numero_ejes','nombre_conductor',
    #     'apellido_conductor','capacidad_camion') \
    #     .annotate(  despachos_realizados=Count('patente_camion'),
    #                 as_integer=Cast('capacidad_camion',IntegerField()),
    #                 suma = Sum('as_integer')) \
    #     .order_by('-despachos_realizados')
    
    # camion_query.aggregate(suma=Sum(as_integer=Cast('capacidad_camion',IntegerField())))
    
    # for camion in camion_queryset:
    #     # result = camion_query.filter(patente_camion=camion['patente_camion']).aggregate(Sum('capacidad_camion'))
    #     print('\n\n\nresult:',camion['available'])
    
    

    workbook = Workbook()
    header_font = Font(name='Calibri', bold=True)
    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Registro de Salida'

    print('Se ha creado la hoja de trabajo')
    # Definir los titulos por columna
    columns = [
        ('Id',5),  #1
        ('Activo',13),
        ('Despachador',12), #28 - falta apellido
        ('RUT Despachador',15), #29
        ('Telefono Despachador Asociado',14), #30
        ('Proyecto',10), #2
        ('Nro. impresiones',10), #3
        ('Nombre cliente',12), #no va?
        ('Rut cliente',15), #no va?
        ('Nombre subcontratista',20), #14
        ('Razón social subcontratista',20), #15
        ('RUT subcontratista',20), #16
        ('Contacto subcontratista',20), #17
        ('Email subcontratista',20), #18
        ('Teléfono subcontratista',20), #19
        ('Nombre conductor',20), #27
        ('Apellido conductor',20), #27
        ('Fecha',11), #4
        ('Hora',8), #5
        ('Patente',8), #20
        ('Marca',7), #21
        ('Modelo',8), #22
        ('Color',8), #23
        ('Volumen',8), #24
        ('Unidad de medida',8), #25
        ('Nro ejes',8), #26
        ('Foto patente',15), #32
        ('Tipo material',13), #13
        ('Punto origen',15), #6
        ('Comuna origen',16), #7
        ('Dirección origen',18), #8
        ('Punto suborigen',15), #9
        ('Punto destino',15), #10
        ('Comuna destino',16), #11
        ('Dirección destino',18), #12
    ]
    row_num = 1
    # Asignar los titulos para cada celda de la cabecera
    for col_num, (column_title, column_width) in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        # set column width
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = column_width
    print('Se han asignado los titulos para cada celda de la cabecera')
    # Iterar por todos los vouchers filtrados por fecha
    for voucher in voucher_queryset:
        row_num += 1

        # serializerOrigen=OrigenSerializer()
        # if Origen.objects.filter(nombre_origen=voucher.nombre_origen).exists():
        #     query_origen = Origen.objects.get(nombre_origen=voucher.nombre_origen)
        #     serializerOrigen = OrigenSerializer(query_origen)
        #     # print('Se serializa el origen')

        # serializerDestino=DestinoSerializer()
        # if Destino.objects.filter(nombre_destino=voucher.punto_destino).exists():
        #     query_destino = Destino.objects.get(nombre_destino=voucher.punto_destino)
        #     serializerDestino = DestinoSerializer(query_destino)
        #     # print('Se serializa el destino')
        
        serializerSubcontratista=SubcontratistaSerializer()
        if Subcontratista.objects.filter(rut=voucher.rut_subcontratista).exists():
            query_subcontratista = Subcontratista.objects.get(rut=voucher.rut_subcontratista)
            serializerSubcontratista = SubcontratistaSerializer(query_subcontratista)
            # print('Se serializa el subcontratista')
            
        # administrador = Proyecto.objects.filter(nombre_origen=despachador.proyecto, is_superuser=True)
        # Define the data for each cell in the row 
        row = [
            voucher.id, #1
            voucher.available,
            voucher.nombre_despachador+' '+voucher.apellido_despachador,#28 - falta apellido
            voucher.rut_despachador, #29
            voucher.telefono_despachador, #30
            voucher.proyecto, #2
            # str(Proyecto.objects.get(id=voucher.id_proyecto)),#2
            voucher.contador_impresiones, #3
            voucher.nombre_cliente, #no va?
            voucher.rut_cliente, #no va?
            voucher.nombre_subcontratista, #14
            voucher.razon_social_subcontratista, #15
            voucher.rut_subcontratista, #16
            voucher.nombre_contacto_subcontratista+' '+voucher.apellido_contacto_subcontratista, #17
            voucher.email_contacto_subcontratista, #18
            voucher.telefono_contacto_subcontratista, #19
            voucher.nombre_conductor, #27
            voucher.apellido_conductor, #27
            voucher.fecha, #4
            voucher.hora, #5
            voucher.patente_camion, #20
            voucher.marca_camion, #21
            voucher.modelo_camion, #22
            voucher.color_camion, #23
            voucher.capacidad_camion, #24
            voucher.unidad_medida, #25
            voucher.numero_ejes, #26
            'https://ohl.faena.app/mediafiles/'+str(voucher.foto_patente), #32
            voucher.tipo_material, #13
            voucher.nombre_origen, #6
            voucher.comuna_origen, #7
            voucher.calle_origen+' '+str(voucher.numero_origen), #8
            voucher.nombre_suborigen, #9
            voucher.nombre_destino, #10
            voucher.comuna_destino, #11
            voucher.calle_destino+' '+str(voucher.numero_destino), #12
        ]
        # print('Se define la data para cada fila')
        # Asignacion de la data para cada celda de la fila
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
        # print('Se asigna la data para cada celda de la fila')


    ### Nueva hoja de trabajo ###
    worksheet = workbook.create_sheet(
        title='Flota Activa',
        index=2,
    )
    print('01')
    # Definir los titulos por columna
    columns = [
        ('Activo',15),
        ('Subcontratista',25), #1
        ('Patente',10), #2
        ('Marca',10), #3
        ('Modelo',10), #4

        ('Color',8), #5
        ('Unidad',8), #6
        ('Numero ejes',13),

        ('Despachos realizados',19), #8
        ('Volumen total desplazado',25), #9
    ]
    print('02')
    row_num = 1
    # Asignar los titulos para cada celda de la cabecera
    for col_num, (column_title, column_width) in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        # set column width
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = column_width

    # Iterar por info de camiones existente en
    for camion_activo in Voucher.objects.raw('SELECT 1 as id, available, nombre_subcontratista, patente_camion, marca_camion, modelo_camion, color_camion, unidad_medida, numero_ejes, count(patente_camion) as despachos,sum(CAST(capacidad_camion as INT)) as volumen_total_desplazado FROM public.api_voucher WHERE available is True GROUP BY available, nombre_subcontratista, patente_camion, marca_camion, modelo_camion, color_camion, unidad_medida, numero_ejes'):
        print('camion_activo: ',camion_activo)
        # camion = Camion.objects.get(patente_camion=camion_activo['patente_camion'])
        row_num += 1
        # Define the data for each cell in the row 
        row = [
            camion_activo.available,
            camion_activo.nombre_subcontratista, #1
            camion_activo.patente_camion, #2
            camion_activo.marca_camion, #3
            camion_activo.modelo_camion, #4

            camion_activo.color_camion, #5
            camion_activo.unidad_medida, #6
            camion_activo.numero_ejes, #7

            camion_activo.despachos, #8
            camion_activo.volumen_total_desplazado, #9
        ]
        # row = [
        #     camion_activo['available'],
        #     camion_activo['nombre_subcontratista'], #1
        #     camion_activo['patente_camion'], #2
        #     camion_activo['marca_camion'], #3
        #     camion_activo['modelo_camion'], #4

        #     camion_activo['color_camion'], #4
        #     camion_activo['unidad_medida'], #6
        #     camion_activo['numero_ejes'], #7

        #     camion_activo['nombre_conductor'], #?
        #     camion_activo['apellido_conductor'], #?
        #     camion_activo['despachos_realizados'], #8
        #     camion_activo['suma'], #9
        # ]
        

        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            
    workbook.save(response)
    return response














@api_view(['GET'])
# @authentication_classes([])
# @permission_classes([])
def exportar_reporte(request,start,hhi,mmi,ssi,end,hhf,mmf,ssf):
    horai=hhi+':'+mmi+':'+ssi
    horaf=hhf+':'+mmf+':'+ssf
    
    if ( not Voucher.objects.filter(fecha__range=(start,end)).exists() ):
        print("\nno hay vouchers")
        res={}
        res = {'request': False, 'error': 'No existen registros para el rango especificado'}
        return Response(res,status=status.HTTP_204_NO_CONTENT)

    response = HttpResponse(
        # content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        content_type='application/vnd.ms-excel',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-reporte.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )

    # voucher_queryset = Voucher.objects.filter(fecha__range=(start,end))
    # print('voucher_queryset t: ', voucher_queryset)

    voucher_inicio = Voucher.objects.filter(fecha=start).filter(hora__range=(horai,'23:59:59')).filter(available=True)
    print('voucher_queryset i: ', voucher_inicio)
    voucher_final = Voucher.objects.filter(fecha=end).filter(hora__range=('00:00:00',horaf)).filter(available=True)
    print('voucher_queryset f: ', voucher_final)
    
    voucher_queryset = voucher_inicio | voucher_final
    print('voucher_queryset R: ', voucher_queryset)

    hoja2_inicio = Voucher.objects.filter(fecha=start).filter(hora__range=(horai,'23:59:59')).filter(available=False)
    print('hoja i: ', voucher_inicio)
    hoja2_final = Voucher.objects.filter(fecha=end).filter(hora__range=('00:00:00',horaf)).filter(available=False)
    print('hoja f: ', voucher_final)
    
    hoja2_queryset = hoja2_inicio | hoja2_final
    print('hoja R: ', voucher_queryset)


    workbook = Workbook()
    header_font = Font(name='Calibri', bold=True)
    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Registro de Salida'

    print('Se ha creado la hoja de trabajo')
    # Definir los titulos por columna
    columns = [
        ('Id',5),  #1
        ('Activo',13), 
        ('Despachador',12), #28 
        ('RUT Despachador',13), #29
        ('Telefono Despachador Asociado',14), #30
        ('Proyecto',10), #2
        ('Nro. impresiones',10), #3
        ('Nombre cliente',12), #no va?
        ('Rut cliente',12), #no va?
        ('Nombre subcontratista',20), #14
        ('Razón social subcontratista',20), #15
        ('RUT subcontratista',20), #16
        ('Contacto subcontratista',20), #17
        ('Email subcontratista',20), #18
        ('Teléfono subcontratista',20), #19
        ('Nombre conductor',20), #27
        ('Apellido conductor',20), #27
        ('Fecha',11), #4
        ('Hora',8), #5
        ('Patente',8), #20
        ('Marca',7), #21
        ('Modelo',8), #22
        ('Color',8), #23
        ('Volumen',8), #24
        ('Unidad de medida',8), #25
        ('Nro ejes',8), #26
        ('Foto patente',15), #32
        ('Tipo material',13), #13
        ('Punto origen',15), #6
        ('Comuna origen',16), #7
        ('Dirección origen',18), #8
        ('Punto suborigen',15), #9
        ('Punto destino',15), #10
        ('Comuna destino',16), #11
        ('Dirección destino',18), #12
    ]
    row_num = 1
    # Asignar los titulos para cada celda de la cabecera
    for col_num, (column_title, column_width) in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        # set column width
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = column_width
    print('Se han asignado los titulos para cada celda de la cabecera')
    # Iterar por todos los vouchers filtrados por fecha
    for voucher in voucher_queryset:
        row_num += 1
        
        # Define the data for each cell in the row 
        row = [
            voucher.id, #1
            voucher.available,
            voucher.nombre_despachador+' '+voucher.apellido_despachador,#28 - falta apellido
            voucher.rut_despachador, #29
            voucher.telefono_despachador, #30
            voucher.proyecto, #2
            # str(Proyecto.objects.get(id=voucher.id_proyecto)),#2
            voucher.contador_impresiones, #3
            voucher.nombre_cliente, #no va?
            voucher.rut_cliente, #no va?
            voucher.nombre_subcontratista, #14
            voucher.razon_social_subcontratista, #15
            voucher.rut_subcontratista, #16
            voucher.nombre_contacto_subcontratista+' '+voucher.apellido_contacto_subcontratista, #17
            voucher.email_contacto_subcontratista, #18
            voucher.telefono_contacto_subcontratista, #19
            voucher.nombre_conductor, #27
            voucher.apellido_conductor, #27
            voucher.fecha, #4
            voucher.hora, #5
            voucher.patente_camion, #20
            voucher.marca_camion, #21
            voucher.modelo_camion, #22
            voucher.color_camion, #23
            voucher.capacidad_camion, #24
            voucher.unidad_medida, #25
            voucher.numero_ejes, #26
            'https://ohl.faena.app/mediafiles/'+str(voucher.foto_patente), #32
            voucher.tipo_material, #13
            voucher.nombre_origen, #6
            voucher.comuna_origen, #7
            voucher.calle_origen+' '+str(voucher.numero_origen), #8
            voucher.nombre_suborigen, #9
            voucher.nombre_destino, #10
            voucher.comuna_destino, #11
            voucher.calle_destino+' '+str(voucher.numero_destino), #12
        ]
        # print('Se define la data para cada fila')
        # Asignacion de la data para cada celda de la fila
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
        # print('Se asigna la data para cada celda de la fila')


    ### Nueva hoja de trabajo ###
    # despachoscamion_queryset = Voucher.objects.filter(fecha__range=(start,end)) \
    #     .values('patente') \
    #     .annotate(despachos_realizados=Count('patente')) \
    #     .order_by('-despachos_realizados')
    # print('serializerV: ', despachoscamion_queryset)
    worksheet = workbook.create_sheet(
        title='Tickets Corregidos',
        index=2,
    )
    print('01')
    # Definir los titulos por columna
    columns = [
        ('Id',5),  #1
        ('Activo',10), 
        ('Despachador',12), #28 
        ('RUT Despachador',13), #29
        ('Telefono Despachador Asociado',14), #30
        ('Proyecto',10), #2
        ('Nro. impresiones',10), #3
        ('Nombre cliente',12), #no va?
        ('Rut cliente',12), #no va?
        ('Nombre subcontratista',20), #14
        ('Razón social subcontratista',20), #15
        ('RUT subcontratista',20), #16
        ('Contacto subcontratista',20), #17
        ('Email subcontratista',20), #18
        ('Teléfono subcontratista',20), #19
        ('Nombre conductor',20), #27
        ('Apellido conductor',20), #27
        ('Fecha',11), #4
        ('Hora',8), #5
        ('Patente',8), #20
        ('Marca',7), #21
        ('Modelo',8), #22
        ('Color',8), #23
        ('Volumen',8), #24
        ('Unidad de medida',8), #25
        ('Nro ejes',8), #26
        ('Foto patente',15), #32
        ('Tipo material',13), #13
        ('Punto origen',15), #6
        ('Comuna origen',16), #7
        ('Dirección origen',18), #8
        ('Punto suborigen',15), #9
        ('Punto destino',15), #10
        ('Comuna destino',16), #11
        ('Dirección destino',18), #12
    ]
    print('02')
    row_num = 1
    # Asignar los titulos para cada celda de la cabecera
    for col_num, (column_title, column_width) in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        # set column width
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = column_width

    # Iterar por info de camiones existente en
    for voucher in hoja2_queryset:
        # camion = Camion.objects.get(patente_camion=camion_activo['patente_camion'])
        row_num += 1
        # Define the data for each cell in the row 
        row = [
            voucher.id, #1
            voucher.available,
            voucher.nombre_despachador+' '+voucher.apellido_despachador,#28 - falta apellido
            voucher.rut_despachador, #29
            voucher.telefono_despachador, #30
            voucher.proyecto, #2
            # str(Proyecto.objects.get(id=voucher.id_proyecto)),#2
            voucher.contador_impresiones, #3
            voucher.nombre_cliente, #no va?
            voucher.rut_cliente, #no va?
            voucher.nombre_subcontratista, #14
            voucher.razon_social_subcontratista, #15
            voucher.rut_subcontratista, #16
            voucher.nombre_contacto_subcontratista+' '+voucher.apellido_contacto_subcontratista, #17
            voucher.email_contacto_subcontratista, #18
            voucher.telefono_contacto_subcontratista, #19
            voucher.nombre_conductor, #27
            voucher.apellido_conductor, #27
            voucher.fecha, #4
            voucher.hora, #5
            voucher.patente_camion, #20
            voucher.marca_camion, #21
            voucher.modelo_camion, #22
            voucher.color_camion, #23
            voucher.capacidad_camion, #24
            voucher.unidad_medida, #25
            voucher.numero_ejes, #26
            'https://ohl.faena.app/mediafiles/'+str(voucher.foto_patente), #32
            voucher.tipo_material, #13
            voucher.nombre_origen, #6
            voucher.comuna_origen, #7
            voucher.calle_origen+' '+str(voucher.numero_origen), #8
            voucher.nombre_suborigen, #9
            voucher.nombre_destino, #10
            voucher.comuna_destino, #11
            voucher.calle_destino+' '+str(voucher.numero_destino), #12
        ]
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            
    workbook.save(response)
    return response